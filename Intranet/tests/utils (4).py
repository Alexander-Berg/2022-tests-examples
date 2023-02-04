import json
import os
import weakref
from collections import Counter
from datetime import datetime, date
import yatest
import vcr
import openpyxl
from django.test.client import Client as BaseClient
from django.utils import timezone


def process_response(response):
    if 'json' in response['body']:
        response['body']['string'] = json.dumps(response['body']['json'])
    return response


def vcr_proxy_matcher(r1, r2):
    """
    Заимствовано отсюда: https://github.yandex-team.ru/easymeeting/easymeeting-api/commit/8644419bfb09f037fd90e6388b281c9c5554adc9#diff-8a66b2566ee3ea0285647f603db9174b

    Костыль для трендбокса + vcr.
    В трендбоксе все https запросы идут через прокси 172.17.0.1:8888.
    Vcr не понимает, на какой хост изначально шёл запрос,
    и не может сматчиться с касетами.
    Поэтому, при запуске через трендбокс,
    вместо стандартной проверки хоста и порта, проверяем,
    что это просто попытка сходить в прокси
    """
    r2_host = r2.host
    r2_port = r2.port

    is_trendbox = os.getenv('TRENDBOX', False)
    if is_trendbox and r1.scheme == 'https' and r1.scheme == r2.scheme:
        r2_host = '172.17.0.1'
        r2_port = 8888

    assert r1.host == r2_host, '{} != {}'.format(r1.host, r2_host)
    assert r1.port == r2_port, '{} != {}'.format(r1.port, r2_port)


def vcr_test(record_mode='none'):  # заменить на 'once', если надо записать новую кассету
    local_path = 'intranet/plan/tests/test_data/vcr_cassettes'
    try:
        path = yatest.common.source_path(local_path)
    except AttributeError:
        path = os.path.join(os.environ["Y_PYTHON_SOURCE_ROOT"], local_path)
    test_vcr_obj = vcr.VCR(
        serializer='json',
        cassette_library_dir=path,
        record_mode=record_mode,
        before_record_response=process_response,
        match_on=(
            'method',
            'scheme',
            'proxy',
            'path',
            'query',
        ),
    )
    test_vcr_obj.register_matcher('proxy', vcr_proxy_matcher)
    return test_vcr_obj


class Client(BaseClient):
    def __init__(self, *args, **kwargs):
        super(Client, self).__init__(*args, **kwargs)
        self.json = JsonClient(weakref.proxy(self))
        self.settings = kwargs['settings']

    def login(self, username):
        self.settings.AUTH_TEST_USER = username
        self.settings.YAUTH_TEST_USER = username


class JsonClient(object):
    """Класс, аналогичный django.test.client.Client, но делающий все запросы с заголовком content_type,
    равным application/json, а для запросов post, put и patch кодирующий тело запроса в JSON.
    Также добавляет к ответу метод json(), который пытается декодировать тело ответа из формата JSON.
    """
    def __init__(self, client):
        self.client = client

    def get(self, path, data={}, **extra):
        return self.jsonify(self.client.get(path, data=data, **extra))

    def post(self, path, data={}, content_type='application/json', **extra):
        return self.jsonify(self.client.post(path, data=json.dumps(data), content_type=content_type, **extra))

    def head(self, path, data={}, **extra):
        return self.jsonify(self.client.head(path, data=data, **extra))

    def options(self, path, data={}, content_type='application/json', **extra):
        return self.jsonify(self.client.options(path, data=json.dumps(data), content_type=content_type, **extra))

    def put(self, path, data={}, content_type='application/json', **extra):
        return self.jsonify(self.client.put(path, data=json.dumps(data), content_type=content_type, **extra))

    def patch(self, path, data={}, content_type='application/json', **extra):
        return self.jsonify(self.client.patch(path, data=json.dumps(data), content_type=content_type, **extra))

    def delete(self, path, data={}, content_type='application/json', **extra):
        return self.jsonify(self.client.delete(path, data=json.dumps(data), content_type=content_type, **extra))

    def jsonify(self, response):
        response.json = lambda: json.loads(response.content) if hasattr(response, 'content') else None
        return response


class Response(object):
    """Эмулятор requests.response"""
    def __init__(self, status_code, content, headers={}, url=None):
        self.status_code = status_code
        self.content = content
        self.headers = headers
        self.url = url
        self.ok = status_code < 400

    def __repr__(self):
        return '<Response [%s]>' % (self.status_code)

    @property
    def text(self):
        """
        Возвращает unicode-объект с телом ответа (из предположения, что кодировка — utf-8).
        """
        return self.content

    def json(self, **kwargs):
        return json.loads(self.text)

    def raise_for_status(self):
        pass


def create_fake_response(content, status_code=200):
    fake_response = Response(status_code, content, '', '')
    return fake_response


def iterables_are_equal(a, b):
    try:
        return Counter(a) == Counter(b)
    except TypeError:  # unhashable type
        b = list(b)
        for x in a:
            try:
                b.remove(x)
            except ValueError:
                return False
        return not b


def list_workdays_some_months(current_day=None, month_count=0):
    # month_count - число следующих месяцев, информацию о рабочих днях которых нужно подгрузить к текущему
    days = [2, 3, 4, 5, 8, 9, 10, 11, 16, 17, 18, 19, 20, 22, 25, 26, 27]

    if current_day is None:
        current_day = timezone.now().date()

    workdays = []
    month_count += 1
    for _ in range(month_count):
        for day in days:
            workdays.append(datetime(current_day.year, current_day.month, day).date())

        next_month = current_day.month + 1 if current_day.month < 12 else 1
        year = current_day.year + 1 if next_month == 1 else current_day.year
        current_day = date(year, next_month, current_day.day)

    return workdays


def _fake_convert_to_html(text, *args, **kwargs):
    return text


class MockIdmResponse:
    def __init__(self, response):
        self._response = response

    @property
    def response(self):
        response = json.dumps(self._response)
        self._response['objects'] = []
        return response


def source_path(path):
    try:
        return yatest.common.source_path(path)
    except AttributeError:
        # only for local pycharm tests
        return os.path.join(os.environ["Y_PYTHON_SOURCE_ROOT"], path)


def compare_xlsx(path_1, path_2):
    wk1 = openpyxl.load_workbook(path_1)
    wk2 = openpyxl.load_workbook(path_2)
    sh1 = wk1.worksheets[0]
    sh2 = wk2.worksheets[0]

    max_row = max(sh1.max_row, sh2.max_row)
    max_column = max(sh1.max_column, sh2.max_column)

    for column in range(1, max_column + 1):
        for row in range(1, max_row + 1):
            value_one = sh1.cell(column=column, row=row).value
            value_two = sh2.cell(column=column, row=row).value
            assert value_one == value_two
