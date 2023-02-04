# coding: utf-8


from collections import defaultdict
import csv
import datetime

import io
from decimal import Decimal

import mock
import openpyxl
import pytest
import xlrd
from django.core.files.uploadedfile import SimpleUploadedFile

from review.core.logic import assemble
from review.core import models
from review.shortcuts import const
from review.lib import serializers
from review.lib import datetimes
from review.core import serializers_extended
from review.core.logic.export import get_properties, get_serialized
from review.frontend.views.export_file import PersonReviewExportView
from tests import helpers


EXPORT_SERIALIZER_FIELDS = serializers_extended.PersonReviewsSerializer.default_fields_ordered


@pytest.mark.parametrize(
    'is_new_scale,mark_at_review', [
        (False, {'mark': 'E', 'text_value': 'E'}),
        (True, {'mark': 'extraordinary', 'text_value': '++++'}),
    ]
)
def test_person_review_export(
    client,
    review_builder,
    person_review_builder,
    review_role_builder,
    finance_builder,
    main_product_builder,
    umbrella_builder,
    umbrella_person_review_builder,
    is_new_scale,
    marks_scale_builder,
    mark_at_review,
):
    if is_new_scale:
        scale = marks_scale_builder()
    else:
        scale = marks_scale_builder(scale={'A': 1, 'B': 1, 'C-': 2, 'C': 3, 'D': 4, 'E': 5})
    review = review_builder(
        mark_mode=const.REVIEW_MODE.MODE_MANUAL,
        goldstar_mode=const.REVIEW_MODE.MODE_MANUAL,
        level_change_mode=const.REVIEW_MODE.MODE_MANUAL,
        salary_change_mode=const.REVIEW_MODE.MODE_AUTO,
        bonus_mode=const.REVIEW_MODE.MODE_AUTO,
        deferred_payment_mode=const.REVIEW_MODE.MODE_AUTO,
        options_rsu_mode=const.REVIEW_MODE.MODE_AUTO,
        scale=scale,
    )
    previous_review = review_builder(
        goldstar_mode=const.REVIEW_MODE.MODE_MANUAL,
        start_date=datetimes.shifted(review.start_date, months=-6),
        finish_date=datetimes.shifted(review.finish_date, months=-6),
        scale=marks_scale_builder(scale={'A': 1, 'B': 1, 'C-': 2, 'C': 3, 'D': 4, 'E': 5}),
    )
    previous_previous_review = review_builder(
        goldstar_mode=const.REVIEW_MODE.MODE_MANUAL,
        start_date=datetimes.shifted(review.start_date, months=-12),
        finish_date=datetimes.shifted(review.finish_date, months=-12),
    )

    superreviewer = review_role_builder(
        review=review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    )
    review_role_builder(
        review=previous_review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
        person=superreviewer.person,
    )
    review_role_builder(
        review=previous_previous_review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
        person=superreviewer.person,
    )
    salary_change = 20
    person_review = person_review_builder(
        review=review,
        mark=mark_at_review['mark'],
        status=const.PERSON_REVIEW_STATUS.EVALUATION,
        goldstar=const.GOLDSTAR.BONUS_ONLY,
        bonus=20,
        bonus_rsu=400,
        deferred_payment=100500,
        salary_change=salary_change,
        salary_change_absolute=120,
        tag_average_mark='fusrodah',
        taken_in_average=True,
    )
    person = person_review.person
    person_review_builder(
        person=person,
        review=previous_review,
        mark='A',
        status=const.PERSON_REVIEW_STATUS.ANNOUNCED,
        goldstar=const.GOLDSTAR.OPTION_AND_BONUS,
    )
    person_review_builder(
        person=person,
        review=previous_previous_review,
        mark='B*',
        status=const.PERSON_REVIEW_STATUS.ANNOUNCED,
        goldstar=const.GOLDSTAR.OPTION_AND_BONUS,
    )
    finance_builder(
        person=person,
        generate_fields=const.OEBS_DATA_TYPES,
    )
    salary = assemble.get_person_review(
        subject=superreviewer.person,
        id=person_review.id,
        fields_requested=(const.FIELDS.SALARY_VALUE,),
    ).salary_value
    person_review.salary_change_absolute = salary * salary_change / 100
    person_review.save()

    def _get_csv_row():
        response = helpers.get(
            client=client,
            path='/frontend/person-reviews/export/csv/',
            request={
                'reviews': [review.id]
            },
            login=superreviewer.person.login,
        )
        assert response.status_code == 200, response.content

        result = io.StringIO(response.content.decode("utf-8"))
        result = list(csv.DictReader(result))
        assert len(result) >= 1

        return result[0]

    row = _get_csv_row()
    assert row['taken_in_average'] == str(True)

    # сперва проверим пустые umbrella и main_product
    assert row['umbrella'] == ''
    assert row['main_product'] == ''

    # а потом добавим их, и проверим все остальное
    main_product = main_product_builder(
        name='main product fizz buzz 12'
    )
    umbrella = umbrella_builder(
        name='umbrella foo bar 42',
        main_product=main_product_builder(),
    )
    person_review.main_product = main_product
    person_review.umbrella = umbrella
    person_review.save()

    row = _get_csv_row()

    pre = assemble.get_person_review(
        subject=superreviewer.person,
        id=person_review.id,
        fields_requested=const.FIELDS.ALL,
    )
    assert set(EXPORT_SERIALIZER_FIELDS) == set(row), row
    assert row['id'] == str(pre.id)
    assert row['login'] == str(pre.person_login)
    assert row['tabled'] == str(pre.flagged)
    assert row['status'] == const.PERSON_REVIEW_STATUS.VERBOSE[
        const.PERSON_REVIEW_STATUS.EVALUATION]
    assert row['grade_before_review'] == str(pre.level)
    assert row['grade_difference'] == str(pre.level_change)
    assert row['grade_after_review'] == str(pre.level + pre.level_change)
    assert row['mark_at_review'] == mark_at_review['text_value']
    assert row['mark_at_review_as_number'] == '1089' if is_new_scale else str(5)
    assert row['mark_at_previous_review'] == 'A'
    # бага СIA-2055
    assert row['mark_at_previous_review_as_number'] == '' if is_new_scale else str(1)
    assert row['mark_before_the_previous_review'] == 'B*'
    assert row['mark_before_the_previous_review_as_number'] == str(2)
    assert row['salary_before_review'] == str(pre.salary_value)
    assert row['salary_change_percentage'] == str(pre.salary_change)
    assert Decimal(row['salary_after_review']) == Decimal(pre.salary_value + pre.salary_change_absolute)
    assert row['salary_change_absolute'] == str(pre.salary_change_absolute)
    assert row['bonus'] == str(pre.bonus_absolute)
    assert row['bonus_payment_percentage'] == str(pre.bonus)
    assert row['extra_payment'] == str(True)
    assert row['extra_option'] == str(False)
    assert row['currency'] == str(pre.salary_currency)
    assert row['city'] in (person.city_name_ru, person.city_name_en)
    assert row['scale'] == str(pre.profession)
    assert row['full_department_name'] == ' \ '.join(pre.person_department_chain_names)
    assert row['fte'] == str(pre.fte) != const.NOT_SET
    assert row['extra_payment_at_previous_review'] == str(True)
    assert row['extra_option_at_previous_review'] == str(True)
    assert row['review_type'] == 'normal'
    assert row['tag_average_mark'] == 'fusrodah'
    assert row['bonus_rsu'] == '400'
    assert row['deferred_payment'] == str(pre.deferred_payment)
    assert row['main_product'] == main_product.name
    assert row['umbrella'] == '%s / %s' % (umbrella.main_product.name, umbrella.name)
    assert row['updated_at'] == pre.updated_at.strftime('%Y-%m-%d %H:%M')


def test_minus_on_previos_review(client, marks_scale_builder, review_builder, person_review_builder, review_role_builder):
    mark_poor = 'poor'
    mark_poor_short = '-'
    mark_good = "good"
    scale = marks_scale_builder(scale={
        mark_poor: {
            "value": 1008,
            "text_value": mark_poor_short,
        },
        mark_good: {
            "value": 1013,
            "text_value": "+",
        },
    })
    prev_review = review_builder(
        start_date=datetime.date(year=2020, month=1, day=1),
        finish_date=datetime.date(year=2020, month=2, day=1),
        evaluation_from_date=datetime.date(year=2020, month=1, day=1),
        evaluation_to_date=datetime.date(year=2020, month=2, day=1),
        scale=scale,
    )
    prev_pr = person_review_builder(
        mark=mark_poor,
        status=const.PERSON_REVIEW_STATUS.ANNOUNCED,
        review=prev_review,
    )
    cur_review = review_builder(
        start_date=datetime.date(year=2021, month=1, day=1),
        finish_date=datetime.date(year=2021, month=2, day=1),
        evaluation_from_date=datetime.date(year=2021, month=1, day=1),
        evaluation_to_date=datetime.date(year=2021, month=2, day=1),
        scale=scale,
    )
    person_review_builder(
        mark=mark_good,
        person=prev_pr.person,
        review=cur_review,
        status=const.PERSON_REVIEW_STATUS.APPROVAL,
    )
    viewer = review_role_builder(
        type=const.ROLE.REVIEW.SUPERREVIEWER,
        review=prev_review,
    ).person
    review_role_builder(
        type=const.ROLE.REVIEW.SUPERREVIEWER,
        review=cur_review,
        person=viewer,
    )

    response = helpers.get(
        client=client,
        path='/frontend/person-reviews/export/csv/',
        request={
            'reviews': [cur_review.id]
        },
        login=viewer.login,
    )
    result = io.StringIO(response.content.decode('utf-8'))
    result = list(csv.DictReader(result))
    row = result[0]

    assert row['mark_at_previous_review'] == mark_poor_short


@pytest.mark.parametrize(
    'role_type,template_type,status_code', [
        [const.ROLE.REVIEW.ACCOMPANYING_HR, const.EXCEL_TEMPLATE_TYPE.SECRET, 200],
        [const.ROLE.REVIEW.ACCOMPANYING_HR, const.EXCEL_TEMPLATE_TYPE.PUBLIC, 200],
        [const.ROLE.REVIEW.SUPERREVIEWER, const.EXCEL_TEMPLATE_TYPE.PUBLIC, 200],
        [const.ROLE.REVIEW.SUPERREVIEWER, const.EXCEL_TEMPLATE_TYPE.SECRET, 404],
    ]
)
def test_person_review_templated_export(
    role_type,
    template_type,
    status_code,
    client,
    for_templated_export,
    person_review,
    review_role_builder,
    excel_template_builder,
):
    review_participant = review_role_builder(
        review=person_review.review,
        type=role_type,
    ).person
    excel_template_builder(
        review=person_review.review,
        template_type=template_type,
    )

    with mock.patch('review.core.logic.assemble._get_file_from_template', return_value=for_templated_export):
        properties = get_properties(
            person_review.review_id,
            template_type,
            review_participant
        )

        serialized = get_serialized(
            data={},
            user=review_participant,
            review_id=person_review.review_id,
            fields=PersonReviewExportView.FIELDS_REQUIRED_FOR_EXPORT_DATA,
            properties=properties
        )

    if not properties.template:
        return

    wb = openpyxl.load_workbook(io.BytesIO(serialized))
    assert wb.get_sheet_by_name('prefilled').cell(row=1, column=1).value == 'correct_test_value'
    assert wb.get_sheet_by_name('reviews').cell(row=2, column=1).value == person_review.id


def test_person_review_export_filename(
    client,
    review_role_builder,
    person_review_builder,
):
    person_review = person_review_builder()
    superreviewer = review_role_builder(
        review=person_review.review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    ).person

    now = datetimes.now()
    with mock.patch('review.lib.datetimes.now', return_value=now):
        response = helpers.get(
            client=client,
            path='/frontend/person-reviews/export/csv/',
            request={
                'reviews': [person_review.review_id]
            },
            login=superreviewer.login,
        )
    assert response.status_code == 200, response.content
    header = response._headers['content-disposition'][1]
    expected_file_name = 'reviews {:%Y-%m-%d %H:%M}.csv'.format(now)
    assert expected_file_name in header, header


@pytest.mark.parametrize('role_type', [const.ROLE.CALIBRATION.ADMIN, const.ROLE.CALIBRATION.CALIBRATOR])
def test_person_review_export_calibration(
    client,
    calibration_builder,
    calibration_person_review_builder,
    calibration_role_builder,
    role_type,
):
    cpr = calibration_person_review_builder(calibration=calibration_builder(status=const.CALIBRATION_STATUS.IN_PROGRESS))
    calibrator = calibration_role_builder(calibration_id=cpr.calibration_id, type=role_type).person
    person_review = cpr.person_review

    response = helpers.get(
        client=client,
        path='/frontend/person-reviews/export/csv/',
        request={
            'reviews': [person_review.review_id]
        },
        login=calibrator.login,
    )
    assert response.status_code == 200, response.content

    result = io.StringIO(response.content.decode('utf-8'))
    result = list(csv.DictReader(result))
    assert result[0]['id'] == str(person_review.id)


TO_UPDATE_PERSON_REVIEW = {
    'mark_at_review': 'excellent',
    'extra_payment': True,
    'extra_option': False,
    'grade_difference': 1,
    'bonus_option_value': 10000000,
    'tag_average_mark': 'fusrodah',
}


@pytest.fixture
def get_test_value_percent():
    return lambda f, pr: dict(
        id=pr.id,
        login=pr.person.login,
        bonus=pr.bonus_absolute,
        bonus_rsu=200,
        deferred_payment=Decimal('100500.01'),
        salary_change_absolute=pr.salary_change_absolute,
        bonus_payment_percentage=50,
        salary_change_percentage=10,
        taken_in_average=True,
        **TO_UPDATE_PERSON_REVIEW
    ).get(f, const.DISABLED)


@pytest.fixture
def get_test_value_absolute():
    return lambda f, pr: dict(
        id=pr.id,
        login=pr.person.login,
        bonus_payment_percentage=200,
        salary_change_percentage=10,
        # this values have to be privileged over *_percentage
        bonus=100,
        salary_change_absolute=20000,
        taken_in_average=False,
        **TO_UPDATE_PERSON_REVIEW
    ).get(f, const.DISABLED)


def test_export_comments(
    client,
    person_review,
    person_review_comment_builder,
    review_role_builder,
):
    texts = ['first', 'second value', 'Lorem, ipsum']
    cmnts = [
        person_review_comment_builder(person_review=person_review, text_wiki=text)
        for text in texts
    ]
    expected_cmnts_str = '\n\n'.join(
        '{} {}:\n{}'.format(it.subject.login, it.created_at.strftime('%d/%m/%y'), it.text_wiki)
        for it in cmnts
    )
    review = person_review.review
    superreviewer = review_role_builder(
        review=review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    )
    response = helpers.get(
        client=client,
        path='/frontend/person-reviews/export-comments/',
        request={
            'reviews': [review.id],
        },
        login=superreviewer.person.login,
    )
    assert response.status_code == 200
    assert response.content
    rb = xlrd.open_workbook(file_contents=response.content)
    sheet = rb.sheet_by_index(0)
    keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]
    result = []
    for row_index in range(1, sheet.nrows):
        row = {
            keys[col_index]: sheet.cell(row_index, col_index).value
            for col_index in range(sheet.ncols)
        }
        result.append(row)
    row = result[0]

    assert row['id'] == person_review.id
    assert row['login'] == person_review.person.login
    assert row['comments'] == expected_cmnts_str


@pytest.mark.parametrize(
    'path,serializer,file_type', [
        ('/frontend/person-reviews/import/', serializers.CSVSerializer, 'csv'),
        ('/frontend/person-reviews/import/', serializers.XLSSerializer, 'xls'),
        ('/frontend/person-reviews/import/csv/', serializers.CSVSerializer, 'csv'),
        ('/frontend/person-reviews/import/xls/', serializers.XLSSerializer, 'xls'),
    ]
)
def test_person_review_import_success(
    client,
    finance_builder,
    get_test_value_percent,
    get_test_value_absolute,
    review_builder,
    person_review_builder,
    review_role_builder,
    path,
    serializer,
    file_type,
):
    review = review_builder(
        mark_mode=const.REVIEW_MODE.MODE_MANUAL,
        goldstar_mode=const.REVIEW_MODE.MODE_MANUAL,
        level_change_mode=const.REVIEW_MODE.MODE_DISABLED,
        salary_change_mode=const.REVIEW_MODE.MODE_MANUAL,
        bonus_mode=const.REVIEW_MODE.MODE_MANUAL,
        options_rsu_mode=const.REVIEW_MODE.MODE_MANUAL,
        deferred_payment_mode=const.REVIEW_MODE.MODE_MANUAL,
    )
    superreviewer = review_role_builder(
        review=review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    )
    person_review_percent = person_review_builder(
        review=review,
        mark='extraordinary',
        goldstar=const.GOLDSTAR.OPTION_AND_BONUS,
        level_change=0,
        salary_change=0,
        bonus=100,
        options_rsu=1000,
        bonus_rsu=300,
        deferred_payment=100,
        taken_in_average=False,
    )

    person_review_abs_bonus = person_review_builder(
        review=review,
        mark='extraordinary',
        goldstar=const.GOLDSTAR.OPTION_AND_BONUS,
        level_change=0,
        salary_change=0,
        bonus=100,
        options_rsu=1000,
        taken_in_average=True,
    )
    finance_builder(
        person=person_review_abs_bonus.person,
        generate_fields=const.OEBS_DATA_TYPES,
    )
    person_review_absolute = assemble.get_person_review(
        subject=superreviewer.person,
        id=person_review_abs_bonus.id,
        fields_requested=const.FIELDS.ALL,
    )

    DISABLED_FIELDS = (
        'grade_difference',
    )
    params_file_content = [
        {
            field: getter(field, pr) if field not in DISABLED_FIELDS else const.DISABLED
            for field in EXPORT_SERIALIZER_FIELDS
        }
        for getter, pr in (
            (get_test_value_percent, person_review_percent),
            (get_test_value_absolute, person_review_abs_bonus)
        )
    ]
    serialized = serializer.serialize(params_file_content)
    as_file = SimpleUploadedFile(
        name='data_file.{}'.format(file_type),
        content=serialized,
        content_type='text/{}'.format(file_type),
    )
    helpers.post_multipart_data(
        client=client,
        path=path,
        request={'data_file': as_file},
        login=superreviewer.person.login,
    )

    updated = helpers.fetch_model(person_review_percent)
    assert updated.mark == get_test_value_percent('mark_at_review', person_review_percent)
    assert updated.goldstar == const.GOLDSTAR.BONUS_ONLY
    assert updated.bonus == get_test_value_percent('bonus_payment_percentage', person_review_percent)
    assert updated.options_rsu == get_test_value_percent('bonus_option_value', person_review_percent)
    assert updated.bonus_rsu == get_test_value_percent('bonus_rsu', person_review_percent)
    assert updated.deferred_payment == get_test_value_percent('deferred_payment', person_review_percent)
    assert updated.salary_change == get_test_value_percent('salary_change_percentage', person_review_percent)
    # not changed — disabled
    assert updated.level_change == person_review_percent.level_change
    assert updated.tag_average_mark == get_test_value_percent('tag_average_mark', person_review_percent)
    assert updated.taken_in_average == get_test_value_percent('taken_in_average', person_review_percent)

    change = updated.changes.get()
    helpers.assert_is_substructure(
        {
            const.FIELDS.MARK: {
                'old': 'extraordinary',
                'new': 'excellent',
            },
            const.FIELDS.GOLDSTAR: {
                'old': const.GOLDSTAR.OPTION_AND_BONUS,
                'new': const.GOLDSTAR.BONUS_ONLY,
            }
        },
        change.diff
    )
    assert change.subject_type == const.PERSON_REVIEW_CHANGE_TYPE.FILE
    assert updated.comments.count() == 0

    updated = helpers.fetch_model(person_review_abs_bonus)
    salary = Decimal(person_review_absolute.salary_value)
    expect_abs_bonus = get_test_value_absolute('bonus', person_review_abs_bonus)
    expect_perc_bonus = Decimal(expect_abs_bonus) / salary * 100
    expect_abs_salary_change = get_test_value_absolute('salary_change_absolute', person_review_abs_bonus)
    expect_perc_salary_change = expect_abs_salary_change / (salary / 100)
    assert updated.bonus_absolute == expect_abs_bonus
    assert updated.bonus == expect_perc_bonus.quantize(Decimal('0.01'))
    assert updated.salary_change_absolute == expect_abs_salary_change
    assert updated.salary_change == expect_perc_salary_change.quantize(Decimal('0.01'))
    assert updated.taken_in_average == get_test_value_absolute('taken_in_average', person_review_abs_bonus)


@pytest.fixture
def gradient_file_builder(
    person_review_builder,
    review_role_builder,
):
    person_review = person_review_builder()
    superreviewer = review_role_builder(
        review=person_review.review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    ).person

    def builder(umbrella=None, main_product=None):
        params = {'id': person_review.id}
        if umbrella:
            params['umbrella'] = umbrella
        if main_product:
            params['main_product'] = main_product
        serialized = serializers.XLSSerializer.serialize([params])
        file_ = SimpleUploadedFile(
            name='data_file.xls',
            content=serialized,
            content_type='text/xls',
        )
        return person_review, superreviewer, file_

    return builder


def test_import_gradient(
    client,
    gradient_file_builder,
    main_product_builder,
    umbrella_builder,
):
    main_product_new = main_product_builder(name='new_mp')
    umbrella_new = umbrella_builder(
        main_product=main_product_new,
        name='new_u',
    )
    person_review, superreviewer, file_ = gradient_file_builder(
        umbrella=' / '.join((main_product_new.name, umbrella_new.name)),
        main_product=main_product_new.name,
    )

    main_product_old = main_product_builder(name='old')
    umbrella_old = umbrella_builder(
        main_product=main_product_old,
        name='old',
    )
    person_review.main_product = main_product_old
    person_review.umbrella = umbrella_old
    person_review.save()

    helpers.post_multipart_data(
        client=client,
        path='/frontend/person-reviews/import/xls/',
        request={'data_file': file_},
        login=superreviewer.login,
    )

    person_review.refresh_from_db()
    assert person_review.umbrella == umbrella_new
    assert person_review.main_product == main_product_new


def test_import_gradient_set_default_umbrella(
    client,
    gradient_file_builder,
    main_product_builder,
    umbrella_builder,
):
    default_umbrella = umbrella_builder()
    person_review, superreviewer, file_ = gradient_file_builder(
        umbrella=default_umbrella.name,
    )

    person_review.umbrella = umbrella_builder(
        main_product=main_product_builder(name='old'),
        name='old',
    )
    person_review.save()

    helpers.post_multipart_data(
        client=client,
        path='/frontend/person-reviews/import/xls/',
        request={'data_file': file_},
        login=superreviewer.login,
    )

    person_review.refresh_from_db()
    assert person_review.umbrella == default_umbrella


def test_import_gradient_send_trash(
    client,
    gradient_file_builder,
    main_product_builder,
    umbrella_builder,
):
    person_review, superreviewer, file_ = gradient_file_builder(
        umbrella='random_name',
        main_product='random_another',
    )

    umbrella = umbrella_builder(name='old')
    main_product = main_product_builder(name='old')
    person_review.umbrella = umbrella
    person_review.main_product = main_product
    person_review.save()

    helpers.post_multipart_data(
        client=client,
        path='/frontend/person-reviews/import/xls/',
        request={'data_file': file_},
        login=superreviewer.login,
    )

    person_review.refresh_from_db()
    assert person_review.umbrella == umbrella
    assert person_review.main_product == main_product


@pytest.mark.parametrize('delimiter', [',', ';'])
def test_csv_dialects(
    client,
    person_review_builder,
    review_role_builder,
    delimiter,
):
    person_review = person_review_builder(tag_average_mark='0')
    superreviewer = review_role_builder(
        review=person_review.review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    ).person
    tag_average_mark = '1'
    file_content = [{
        'id': person_review.id,
        'tag_average_mark': tag_average_mark,
    }]
    serialized = serializers.CSVSerializer.serialize(
        file_content,
        delimiter=delimiter,
    )
    file_ = SimpleUploadedFile(
        name='data_file.xls',
        content=serialized,
        content_type='text/xls',
    )

    helpers.post_multipart_data(
        client=client,
        path='/frontend/person-reviews/import/xls/',
        request={'data_file': file_},
        login=superreviewer.login,
    )

    person_review.refresh_from_db()
    assert person_review.tag_average_mark == tag_average_mark


@pytest.mark.parametrize(
    'path,serializer,file_type', [
        ('/frontend/person-reviews/import/', serializers.CSVSerializer, 'csv'),
        ('/frontend/person-reviews/import/', serializers.XLSSerializer, 'xls'),
        ('/frontend/person-reviews/import/csv/', serializers.CSVSerializer, 'csv'),
        ('/frontend/person-reviews/import/xls/', serializers.XLSSerializer, 'xls'),
    ]
)
def test_person_review_import_success_no_marks(
    client,
    finance_builder,
    get_test_value_absolute,
    review_builder,
    person_review_builder,
    review_role_builder,
    path,
    serializer,
    file_type,
):
    review = review_builder(
        mark_mode=const.REVIEW_MODE.MODE_DISABLED,
        goldstar_mode=const.REVIEW_MODE.MODE_MANUAL,
        level_change_mode=const.REVIEW_MODE.MODE_DISABLED,
        salary_change_mode=const.REVIEW_MODE.MODE_DISABLED,
        bonus_mode=const.REVIEW_MODE.MODE_MANUAL,
        options_rsu_mode=const.REVIEW_MODE.MODE_MANUAL,
    )
    superreviewer = review_role_builder(
        review=review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    )

    person_review_abs_bonus = person_review_builder(
        review=review,
        goldstar=const.GOLDSTAR.OPTION_AND_BONUS,
        level_change=0,
        salary_change=0,
        bonus=100,
        options_rsu=1000,
    )
    finance_builder(
        person=person_review_abs_bonus.person,
        generate_fields=const.OEBS_DATA_TYPES,
    )
    pr_obj = assemble.get_person_review(
        subject=superreviewer.person,
        id=person_review_abs_bonus.id,
        fields_requested=const.FIELDS.ALL,
    )

    disabled_fields = (
        'salary_change_percentage',
        'grade_difference',
    )
    content = dict.fromkeys(disabled_fields, const.DISABLED)
    for field in set(EXPORT_SERIALIZER_FIELDS) - set(disabled_fields):
        content[field] = get_test_value_absolute(field, person_review_abs_bonus)
    content['mark_at_review'] = None
    serialized = serializer.serialize([content])
    as_file = SimpleUploadedFile(
        name='data_file.{}'.format(file_type),
        content=serialized,
        content_type='text/{}'.format(file_type),
    )

    helpers.post_multipart_data(
        client=client,
        path=path,
        request={'data_file': as_file},
        login=superreviewer.person.login,
    )

    updated = helpers.fetch_model(person_review_abs_bonus)
    assert updated.mark == const.MARK.DEFAULT


def test_person_review_import_format(
    client,
    get_test_value_percent,
    person_review_builder,
    review_builder,
    review_role_builder,
):
    review = review_builder(
        mark_mode=const.REVIEW_MODE.MODE_MANUAL,
        goldstar_mode=const.REVIEW_MODE.MODE_MANUAL,
    )
    superreviewer = review_role_builder(
        review=review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    )
    person_reviews = [
        person_review_builder(
            review=review,
            mark='extraordinary',
        )
        for _ in range(2)
    ]

    params_file_content = [
        {
            field: get_test_value_percent(field, r)
            for field in EXPORT_SERIALIZER_FIELDS
        }
        for r in person_reviews
    ]
    params_file_content[0]['mark_at_review'] = 'U'
    params_file_content[1]['mark_at_review'] = None
    params_file_content[1]['extra_option'] = None
    serialized = serializers.CSVSerializer.serialize(params_file_content)
    as_file = SimpleUploadedFile(
        name='data_file.csv',
        content=serialized,
        content_type='text/csv',
    )
    response = helpers.post_multipart_data(
        client=client,
        path='/frontend/person-reviews/import/',
        request={'data_file': as_file},
        login=superreviewer.person.login,
        expect_status=400,
    )
    error = response['errors']['__all__']
    assert error['code'] == 'FILE_INCORRECT'
    err_params = error['params']
    assert all(col in err_params for col in ('mark_at_review', 'extra_option'))
    assert {'0', '1'} == set(err_params['mark_at_review'].keys())
    assert {'1'} == set(err_params['extra_option'].keys())


@pytest.mark.parametrize(
    'path,serializer,file_type', [
        ('/frontend/person-reviews/import/', serializers.CSVSerializer, 'csv'),
        ('/frontend/person-reviews/import/', serializers.XLSSerializer, 'xls'),
        ('/frontend/person-reviews/import/csv/', serializers.CSVSerializer, 'csv'),
        ('/frontend/person-reviews/import/xls/', serializers.XLSSerializer, 'xls'),
    ]
)
def test_person_review_import_failed_rollback(
    client,
    get_test_value_percent,
    review_builder,
    person_review_builder,
    person_review_role_builder,
    path,
    serializer,
    file_type,
):
    review = review_builder(
        mark_mode=const.REVIEW_MODE.MODE_MANUAL,
        goldstar_mode=const.REVIEW_MODE.MODE_MANUAL,
    )
    person_review_available = person_review_builder(
        review=review,
        mark='extraordinary',
    )
    person_review_unavailable = person_review_builder(
        review=review,
        mark='extraordinary',
    )
    reviewer = person_review_role_builder(
        person_review=person_review_available,
        type=const.ROLE.PERSON_REVIEW.TOP_REVIEWER,
    )

    params_file_content = [
        {
            field: get_test_value_percent(field, r)
            for field in EXPORT_SERIALIZER_FIELDS
        }
        for r in [
            person_review_available,
            person_review_unavailable
        ]
    ]
    serialized = serializer.serialize(params_file_content)
    as_file = SimpleUploadedFile(
        name='data_file.{}'.format(file_type),
        content=serialized,
        content_type='text/{}'.format(file_type),
    )

    helpers.post_multipart_data(
        client=client,
        path=path,
        request={'data_file': as_file},
        login=reviewer.person.login,
        expect_status=400
    )

    person_review_available = helpers.fetch_model(person_review_available)
    assert person_review_available.mark == 'extraordinary'
    assert person_review_available.changes.count() == 0
    person_review_available = helpers.fetch_model(person_review_unavailable)
    assert person_review_available.mark == 'extraordinary'
    assert person_review_available.changes.count() == 0


def test_wrong_file_format(
    client,
    bad_file,
    person,
):
    response = helpers.post_multipart_data(
        client=client,
        path='/frontend/person-reviews/import/',
        request={'data_file': bad_file},
        login=person.login,
        expect_status=400,
    )
    assert response['errors']['__all__']['code'] == 'CANT_PARSE_FILE'


def test_import_comments(
    client,
    person_review_builder,
    review_role_builder,
    robot,
):
    person_review = person_review_builder()
    superreviewer = review_role_builder(
        review=person_review.review,
        type=const.ROLE.REVIEW.SUPERREVIEWER,
    )
    tmpl = 'Сотруднику выдана ненастоящая премия в размере {special_bonus} {currency}. Деоать с ними можно {additional_info}'
    bonus = 12345
    currency = 'Дырок от бублика'
    additional_info = 'whateveruwish'
    content = [{'id': person_review.id, 'special_bonus': bonus, 'currency': currency, 'additional_info': additional_info}]
    as_file = SimpleUploadedFile(
        name='data_file.csv',
        content=serializers.CSVSerializer.serialize(content),
        content_type='text/csv',
    )

    helpers.post_multipart_data(
        client=client,
        path='/frontend/import/comments/',
        request={'data_file': as_file, 'text_tmpl': tmpl},
        login=superreviewer.person.login,
    )

    updated = helpers.fetch_model(person_review)
    comment = updated.comments.first()
    assert comment
    row = content[0]
    row.pop('id')
    assert comment.text_wiki == tmpl.format(**row)
    assert comment.subject == robot


def test_import_comments_view(client):
    response = helpers.get(
        client=client,
        path='/frontend/import/',
    )
    assert response.status_code == 200
