from io import BytesIO

from django.core.urlresolvers import reverse
from openpyxl import load_workbook


def test_export_control_plan_success(db, client, control_plan):
    url = reverse('api_v1:export', kwargs={'obj_class': 'controlplan'})
    response = client.get(url, {'obj_pks': control_plan.id})
    assert response.status_code == 200
    expected = ('Content-Disposition', 'attachment; filename=controlplan.xlsx')
    assert response._headers['content-disposition'] == expected

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 2
    assert ws['A2'].value == str(control_plan.control)


def test_export_many_control_plan_success(db, client, control_plan, control_plan_two):
    url = reverse('api_v1:export', kwargs={'obj_class': 'controlplan'})
    response = client.get(url, {'obj_pks': '{},{}'.format(control_plan.id, control_plan_two.id)})
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 3


def test_export_many_control_plan_queries_success(db, client, django_assert_num_queries,
                                                  default_queries_count, control_plan,
                                                  control_plan_two):
    url = reverse('api_v1:export', kwargs={'obj_class': 'controlplan'})
    with django_assert_num_queries(default_queries_count + 12):
        response = client.get(
            path=url,
            data={'obj_pks': '{},{}'.format(control_plan.id, control_plan_two.id)},
        )
    assert response.status_code == 200


def test_export_all_control_plan_success(db, client, control_plan, control_plan_two):
    url = reverse('api_v1:export', kwargs={'obj_class': 'controlplan'})
    response = client.get(url)
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 3


def test_export_filter_by_control_plan_success(db, client, process, control_plan, control_plan_two):
    url = reverse('api_v1:export', kwargs={'obj_class': 'controlplan'})
    response = client.get(url, {'filter_by': 'process:{}'.format(process.id)})
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 2


def test_export_ipe_success(db, client, ipe):
    url = reverse('api_v1:export', kwargs={'obj_class': 'ipe'})
    response = client.get(url, {'obj_pks': ipe.id})
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 2


def test_export_controltest_success(db, client, control_test):
    url = reverse('api_v1:export', kwargs={'obj_class': 'controltest'})
    response = client.get(url, {'obj_pks': control_test.id})
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 2


def test_export_deficiency_success(db, client, deficiency):
    url = reverse('api_v1:export', kwargs={'obj_class': 'deficiency'})
    response = client.get(url, {'obj_pks': deficiency.id})
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 2


def test_export_deficiency_group_success(db, client, deficiency_group):
    url = reverse('api_v1:export', kwargs={'obj_class': 'deficiencygroup'})
    response = client.get(url, {'obj_pks': deficiency_group.id})
    assert response.status_code == 200

    wb = load_workbook(filename=BytesIO(response.content))
    ws = wb.active
    assert ws.max_row == 2
